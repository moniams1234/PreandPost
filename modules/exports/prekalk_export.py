"""prekalk_export.py — XLSX builder for PreKalkulacja."""
from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook

from modules.exports.xlsx_styles import xlsx_write_sheet, apply_alert_formatting
from modules.calculations.summaries import build_prekalk_summary


def export_prekalkulacja_xlsx(
    df_pk: pd.DataFrame,
    df_orders: pd.DataFrame | None,
    df_tektura: pd.DataFrame | None,
    df_material: pd.DataFrame | None,
    df_tools: pd.DataFrame | None,
    df_wydajnosc: pd.DataFrame | None,
    df_maszyny: pd.DataFrame | None,
    tpm_thr: float = 60.0,
    cm_thr: float = 40.0,
) -> io.BytesIO:
    """Build prekalkulacja.xlsx and return BytesIO."""
    wb = Workbook()
    for ws in wb.worksheets:
        wb.remove(ws)

    pct_pk = {"TPM%", "CM%"}
    curr_pk = {
        "Sales Value", "TPM", "CM", "Total DL", "Total Materials Cost",
        "Koszt tektury", "Koszt gilotyny", "Pliki final", "Offset inks",
        "Płyty offset", "Opakowania zbiorcze", "Klej", "Lakiery", "Other Materials",
        "Prepress",
    } | {c for c in df_pk.columns if c.startswith("Koszt ")}

    # ── prekalkulacja ─────────────────────────────────────────────────────────
    ws_pk = wb.create_sheet("prekalkulacja")
    xlsx_write_sheet(ws_pk, df_pk,
                     curr_cols=curr_pk, pct_cols=pct_pk,
                     date_cols={"Data faktury"}, ym_cols={"Miesiąc faktury"})

    # ── orders ────────────────────────────────────────────────────────────────
    if df_orders is not None and not df_orders.empty:
        xlsx_write_sheet(wb.create_sheet("orders"), df_orders)

    # ── tektura ───────────────────────────────────────────────────────────────
    if df_tektura is not None and not df_tektura.empty:
        xlsx_write_sheet(wb.create_sheet("tektura"), df_tektura,
                         curr_cols={"(m.value/ m.qty)", "Cena tektury (m.value/m.qty)"})

    # ── usługa na surowcu ─────────────────────────────────────────────────────
    if df_material is not None and not df_material.empty:
        xlsx_write_sheet(wb.create_sheet("usługa na surowcu"), df_material)

    # ── tools ─────────────────────────────────────────────────────────────────
    if df_tools is not None and not df_tools.empty:
        xlsx_write_sheet(wb.create_sheet("tools"), df_tools)

    # ── wydajność ─────────────────────────────────────────────────────────────
    if df_wydajnosc is not None and not df_wydajnosc.empty:
        xlsx_write_sheet(wb.create_sheet("wydajność"), df_wydajnosc,
                         curr_cols={"Stawka rbg"})

    # ── maszyny ───────────────────────────────────────────────────────────────
    if df_maszyny is not None and not df_maszyny.empty:
        xlsx_write_sheet(wb.create_sheet("maszyny"), df_maszyny)

    # ── podsumowanie ──────────────────────────────────────────────────────────
    df_sum = build_prekalk_summary(df_pk)
    ws_s = wb.create_sheet("podsumowanie")
    if not df_sum.empty:
        xlsx_write_sheet(ws_s, df_sum,
                         curr_cols={"Sales Value", "TPM", "CM"},
                         pct_cols={"TPM%", "CM%"})
        apply_alert_formatting(ws_s, df_sum, "TPM%", "CM%", tpm_thr, cm_thr)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
