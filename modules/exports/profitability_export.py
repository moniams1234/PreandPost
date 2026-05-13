"""profitability_export.py — XLSX builder for PostKalkulacja Profitability."""
from __future__ import annotations

import io
import pandas as pd
from openpyxl import Workbook

from modules.exports.xlsx_styles import (
    xlsx_write_sheet, apply_alert_formatting,
    FILL_DL_H, FILL_DL_R, FILL_MT_H, FILL_MT_R, FILL_HDR,
)
from modules.utils.formatting import is_currency_col

HIDDEN_COLS: set[str] = {
    "Kontrbigi [29]", "Duplex [30]", "Offset [31]",
    "Add on evey box [32]", "Autobottom / ekstra lim [33]",
    "Pantone [34]", "E-flute [35]", "Item number inlay [36]",
    "POD price [37]", "Clames [38]", "Transport [39]",
    "Kooperacja [40]", "B2 price [41]", "B1 price [42]",
    "Energia [43]", "Fix price [44]", "Cena TKW [45]",
    "Click price [46]",
    "Płyty lakierujące [21]", "Matryca Braille- Grawer [22]",
    "Patryca Braille- Wewnatrz [23]",
}

MAT_NAMES_STD = [
    "Papier [16]", "Klej [17]", "Lakiery [20]",
    "Opakowania zbiorcze [24]", "Other Materials",
    "Offset inks", "Płyta offsetowa", "Kliki final", "Total Materials",
]


def build_xlsx(
    result: dict,
    rates: dict,
    click_costs: dict,
    prepress: dict,
    other_pct: float,
    tpm_thr: float,
    cm_thr: float,
) -> io.BytesIO:
    """Build full PostKalkulacja XLSX and return BytesIO."""
    from modules.calculations.profitability_engine import exclude_oticon_zam_rows

    wb = Workbook()
    for ws in wb.worksheets:
        wb.remove(ws)

    df_prof = result["df_prof"]
    df_prof_report = exclude_oticon_zam_rows(df_prof, result.get("klient_col"))
    df_czasy = result["df_czasy"]
    df_kliki = result["df_kliki"]
    df_farby_pivot = result["df_farby_pivot"]
    machine_cols = result["machine_cols"]
    klient_col = result["klient_col"]
    grp_col = klient_col or "Zlecenie produkcyjne"

    # New supplementary sheets from result dict
    df_material = result.get("df_material")
    df_tektura_aug = result.get("df_tektura")
    df_orders = result.get("df_orders")

    DL_SET = set(machine_cols + ["Prepress costs", "koszt gilotyny", "Total DL"])
    MAT_SET = {c for c in df_prof.columns
               if any(" ".join(m.split()).lower() in " ".join(c.split()).lower()
                      for m in MAT_NAMES_STD)}
    curr_c = {c for c in df_prof.columns if is_currency_col(c)}

    # ── Profitability ─────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Profitability")
    xlsx_write_sheet(
        ws_p, df_prof,
        dl_set=DL_SET, mat_set=MAT_SET,
        hidden=HIDDEN_COLS,
        curr_cols=curr_c - {"TPM%", "CM%"},
        pct_cols={"TPM%", "CM%"},
        date_cols={"Data faktury"},
        ym_cols={"Miesiąc faktury"},
    )

    # ── zlecenia wg Lini faktury ──────────────────────────────────────────────
    source_col = "_matched_source" if "_matched_source" in df_prof.columns else (
        "Źródło Sales Value" if "Źródło Sales Value" in df_prof.columns else None
    )
    if source_col:
        src_series = df_prof[source_col].fillna("").astype(str).str.lower()
        df_linie_src = df_prof[src_series.str.contains("faktury linie", na=False)].copy()
    else:
        df_linie_src = pd.DataFrame(columns=df_prof.columns)

    ws_linie_src = wb.create_sheet("zlecenia wg Lini faktury")
    xlsx_write_sheet(
        ws_linie_src,
        df_linie_src,
        dl_set=DL_SET,
        mat_set=MAT_SET,
        hidden=HIDDEN_COLS,
        curr_cols={c for c in df_linie_src.columns if is_currency_col(c)} - {"TPM%", "CM%"},
        pct_cols={"TPM%", "CM%"},
        date_cols={"Data faktury"},
        ym_cols={"Miesiąc faktury"},
    )

    # ── usługa na surowcu ─────────────────────────────────────────────────────
    if df_material is not None and not df_material.empty:
        ws_mat = wb.create_sheet("usługa na surowcu")
        xlsx_write_sheet(ws_mat, df_material)

    # ── tektura (with liczba cięć column) ────────────────────────────────────
    if df_tektura_aug is not None and not df_tektura_aug.empty:
        ws_tek = wb.create_sheet("tektura")
        xlsx_write_sheet(ws_tek, df_tektura_aug)

    # ── orders ────────────────────────────────────────────────────────────────
    if df_orders is not None and not df_orders.empty:
        ws_orders = wb.create_sheet("orders")
        xlsx_write_sheet(ws_orders, df_orders)

    # ── czasy ─────────────────────────────────────────────────────────────────
    if df_czasy is not None:
        ws_c = wb.create_sheet("czasy")
        df_c = df_czasy.drop(columns=["_rate"], errors="ignore")
        xlsx_write_sheet(ws_c, df_c,
                         curr_cols={c for c in df_c.columns if is_currency_col(c)})

    # ── Kliki ─────────────────────────────────────────────────────────────────
    if df_kliki is not None:
        ws_k = wb.create_sheet("Kliki")
        xlsx_write_sheet(ws_k, df_kliki,
                         curr_cols={c for c in df_kliki.columns if is_currency_col(c)})

    # ── Farby Offset ──────────────────────────────────────────────────────────
    if df_farby_pivot is not None:
        ws_f = wb.create_sheet("Farby Offset")
        xlsx_write_sheet(ws_f, df_farby_pivot,
                         curr_cols={c for c in df_farby_pivot.columns if is_currency_col(c)})

    # ── Stawki ────────────────────────────────────────────────────────────────
    ws_st = wb.create_sheet("Stawki")
    df_st = pd.DataFrame(list(rates.items()), columns=["Nazwa maszyny", "Stawka rbg (PLN/h)"])
    xlsx_write_sheet(ws_st, df_st, curr_cols={"Stawka rbg (PLN/h)"})

    # ── Koszty klików ─────────────────────────────────────────────────────────
    ws_kk = wb.create_sheet("Koszty klików")
    cc_rows = [
        {"Maszyna": k, "Kolor": c, "Koszt PLN/sep": v}
        for k, colors in click_costs.items()
        for c, v in colors.items()
    ]
    df_kk = (pd.DataFrame(cc_rows) if cc_rows
              else pd.DataFrame(columns=["Maszyna", "Kolor", "Koszt PLN/sep"]))
    xlsx_write_sheet(ws_kk, df_kk, curr_cols={"Koszt PLN/sep"})

    # ── Prepress ──────────────────────────────────────────────────────────────
    pp_rows = [{"Klient": k, "Digital": v["digital"], "Offset": v["offset"]}
               for k, v in prepress.items()]
    if not pp_rows:
        pp_rows = [{"Klient": "(domyślna)", "Digital": 10.0, "Offset": 40.0}]
    ws_pp = wb.create_sheet("Prepress")
    xlsx_write_sheet(ws_pp, pd.DataFrame(pp_rows), curr_cols={"Digital", "Offset"})

    # ── Parametry ─────────────────────────────────────────────────────────────
    ws_par = wb.create_sheet("Parametry")
    df_par = pd.DataFrame([
        {"Parametr": "Other costs %",     "Wartość": other_pct / 100},
        {"Parametr": "Próg alertu TPM %", "Wartość": tpm_thr / 100},
        {"Parametr": "Próg alertu CM %",  "Wartość": cm_thr / 100},
    ])
    xlsx_write_sheet(ws_par, df_par, pct_cols={"Wartość"})

    # ── Per-month Podsumowanie + Batch ────────────────────────────────────────
    months = sorted(df_prof_report["Miesiąc faktury"].dropna().unique())
    for month in months:
        df_m = df_prof_report[df_prof_report["Miesiąc faktury"] == month].copy()
        if grp_col not in df_m.columns:
            continue
        df_m[grp_col] = df_m[grp_col].fillna("(brak)")
        rows_s = []
        for kl, grp in df_m.groupby(grp_col):
            sv = grp["Sales Value"].sum()
            tpm = grp["TPM"].sum()
            cm = grp["CM"].sum()
            rows_s.append({
                "Klient": kl, "Miesiąc": month,
                "Suma sprzedaży": sv, "Suma TPM": tpm,
                "TPM %": tpm / sv if sv else 0,
                "Suma CM": cm,
                "CM %": cm / sv if sv else 0,
                "Zamówień": len(grp),
                "Digital": (grp["Digital/Offset"] == "Digital").sum()
                if "Digital/Offset" in grp.columns else 0,
                "Offset": (grp["Digital/Offset"] == "Offset").sum()
                if "Digital/Offset" in grp.columns else 0,
                "No printing": (grp["Digital/Offset"] == "no printing").sum()
                if "Digital/Offset" in grp.columns else 0,
            })
        df_s = pd.DataFrame(rows_s)
        safe_m = month.replace("/", "-")
        ws_s = wb.create_sheet(f"Podsum. {safe_m}")
        xlsx_write_sheet(ws_s, df_s,
                         curr_cols={"Suma sprzedaży", "Suma TPM", "Suma CM"},
                         pct_cols={"TPM %", "CM %"})
        apply_alert_formatting(ws_s, df_s, "TPM %", "CM %", tpm_thr, cm_thr)

        if "Batch" in df_m.columns:
            ws_b = wb.create_sheet(f"Batch {safe_m}")
            df_b = (df_m.groupby([grp_col, "Batch"])
                    .size().reset_index(name="Liczba zamówień"))
            xlsx_write_sheet(ws_b, df_b)

    # ── Kokpit ────────────────────────────────────────────────────────────────
    ws_kok = wb.create_sheet("Kokpit")
    kok_rows = []
    for month in months:
        df_m = df_prof_report[df_prof_report["Miesiąc faktury"] == month]
        sv = df_m["Sales Value"].sum()
        tpm = df_m["TPM"].sum()
        cm = df_m["CM"].sum()
        kok_rows.append({
            "Miesiąc": month, "Sprzedaż": sv,
            "TPM": tpm, "TPM %": tpm / sv if sv else 0,
            "CM": cm,  "CM %": cm / sv if sv else 0,
            "Klientów": df_m[grp_col].nunique() if grp_col in df_m.columns else 0,
            "Zamówień": len(df_m),
            "Digital": (df_m["Digital/Offset"] == "Digital").sum()
            if "Digital/Offset" in df_m.columns else 0,
            "Offset": (df_m["Digital/Offset"] == "Offset").sum()
            if "Digital/Offset" in df_m.columns else 0,
            "No printing": (df_m["Digital/Offset"] == "no printing").sum()
            if "Digital/Offset" in df_m.columns else 0,
        })
    if kok_rows:
        xlsx_write_sheet(ws_kok, pd.DataFrame(kok_rows),
                         curr_cols={"Sprzedaż", "TPM", "CM"},
                         pct_cols={"TPM %", "CM %"})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
