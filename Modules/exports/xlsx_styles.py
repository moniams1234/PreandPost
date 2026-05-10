"""xlsx_styles.py — shared openpyxl styles and the universal sheet-writer."""
from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Fills ─────────────────────────────────────────────────────────────────────
FILL_HDR  = PatternFill("solid", start_color="3A0000")   # dark burgundy
FILL_DL_H = PatternFill("solid", start_color="1565A0")   # blue DL header
FILL_DL_R = PatternFill("solid", start_color="D6ECF8")   # blue DL row
FILL_MT_H = PatternFill("solid", start_color="1A7A45")   # green Materials header
FILL_MT_R = PatternFill("solid", start_color="D4F0DF")   # green Materials row
FILL_RED  = PatternFill("solid", start_color="C00000")   # alert red

# ── Fonts ─────────────────────────────────────────────────────────────────────
FONT_WH  = Font(bold=True,  color="FFFFFF", name="Arial", size=9)
FONT_BD  = Font(bold=True,  name="Arial", size=9)
FONT_REG = Font(name="Arial", size=9)
FONT_WH_BD_RED = Font(bold=True, color="FFFFFF", name="Arial", size=9)

# ── Other ─────────────────────────────────────────────────────────────────────
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_R = Alignment(horizontal="right",  vertical="center")
THIN_S  = Side(style="thin", color="D4D4D4")
BORDER  = Border(left=THIN_S, right=THIN_S, top=THIN_S, bottom=THIN_S)

FMT_EUR  = "#,##0.00"
FMT_PCT  = "0.0%"
FMT_DATE = "YYYY-MM-DD"
FMT_YM   = "@"
FMT_INT  = "#,##0"


def _col_group(col: str, dl_set: set, mat_set: set):
    if col in dl_set:
        return FILL_DL_H, FILL_DL_R
    if col in mat_set:
        return FILL_MT_H, FILL_MT_R
    return FILL_HDR, None


def xlsx_write_sheet(
    ws,
    df: pd.DataFrame,
    dl_set: set = None,
    mat_set: set = None,
    hidden: set = None,
    curr_cols: set = None,
    pct_cols: set = None,
    date_cols: set = None,
    ym_cols: set = None,
    int_cols: set = None,
) -> None:
    """Write a DataFrame to an openpyxl worksheet with full professional formatting."""
    dl_set = dl_set or set()
    mat_set = mat_set or set()
    hidden_n = {" ".join(h.split()).lower() for h in (hidden or set())}
    cols = list(df.columns)
    n_rows = len(df)

    # Header
    for ci, col in enumerate(cols, 1):
        hfill, _ = _col_group(col, dl_set, mat_set)
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hfill
        cell.font = FONT_WH
        cell.alignment = ALIGN_C
        cell.border = BORDER

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            _, rfill = _col_group(col, dl_set, mat_set)
            val = row[col]
            if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = FONT_REG
            cell.border = BORDER
            if rfill:
                cell.fill = rfill
            if pct_cols and col in pct_cols:
                cell.number_format = FMT_PCT
                cell.alignment = ALIGN_R
            elif date_cols and col in date_cols:
                cell.number_format = FMT_DATE
                cell.alignment = ALIGN_C
            elif ym_cols and col in ym_cols:
                cell.number_format = FMT_YM
                cell.alignment = ALIGN_C
            elif int_cols and col in int_cols:
                cell.number_format = FMT_INT
                cell.alignment = ALIGN_R
            elif curr_cols and col in curr_cols:
                cell.number_format = FMT_EUR
                cell.alignment = ALIGN_R
            else:
                cell.alignment = ALIGN_L

    # Column widths
    for ci, col in enumerate(cols, 1):
        sample = [str(df.iat[r, ci - 1])[:40] for r in range(min(n_rows, 80))]
        w = max(len(str(col)), max((len(s) for s in sample), default=0)) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w, 8), 44)

    # Hidden columns
    for ci, col in enumerate(cols, 1):
        if " ".join(col.split()).lower() in hidden_n:
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    # Freeze + filter
    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def apply_alert_formatting(
    ws,
    df: pd.DataFrame,
    tpm_col: str = "TPM %",
    cm_col: str = "CM %",
    tpm_thr: float = 60.0,
    cm_thr: float = 40.0,
) -> None:
    """Colour TPM% and CM% cells red when below threshold."""
    if df is None or df.empty:
        return
    thresholds = {tpm_col: tpm_thr / 100.0, cm_col: cm_thr / 100.0}
    for col_name, thr in thresholds.items():
        if col_name not in df.columns:
            continue
        ci = list(df.columns).index(col_name) + 1
        for ri in range(2, len(df) + 2):
            cell = ws.cell(row=ri, column=ci)
            try:
                v = float(cell.value or 0)
            except (TypeError, ValueError):
                continue
            if v < thr:
                cell.fill = FILL_RED
                cell.font = FONT_WH_BD_RED
